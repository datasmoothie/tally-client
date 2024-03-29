import pandas as pd
import os
import tally
import pytest
import openpyxl

def test_add_table(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='client A', default_dataset=ds)

    sheet = build.add_sheet(banner=['gender', 'locality'])

    sheet.options.set_base_position('outside')
    sheet.options.set_answer_format('base', {"font_color":"#F15A30", "bold":True})
    sheet.options.set_question_format('percentage', {"bold":True})
    sheet.options.set_column_format_for_type('base', 1, {"bold":True})
    sheet.options.set_column_format_for_type('percentage', 1, {"bold":True})

    # base can be outside, above (default), hide
    sheet.add_table(stub={'x':'q2b', 'ci':['c%', 'counts'], 'xtotal':True, 'stats':['mean', 'stddev']}, dataset=ds)
    sheet.add_table(stub={'x':'q2b', 'ci':['c%'], 'f':{'locality':[1]}, 'w':'weight_a', 'xtotal':True, 'stats':['mean', 'stddev']}, dataset=ds)
    sheet.add_table(stub={'x':'q1', 'ci':['c%'], 'f':{'locality':[2]}, 'w':'weight_a', 'base':'both', 'xtotal':True}, 
                          options={'row_format':{'rows':[4,5,6], 'format':{'bg_color':'F15A30'}}})
    sheet.add_table(stub={'x':'q3', 'ci':['c%'], 'f':{'locality':[1]}, 'w':'weight_a', 'xtotal':True}, options={'base':'hide'})
    sheet.add_table(stub={'x':'q2b', 'ci':['c%'], 'f':{'locality':[1]}, 'w':'weight_a', 'xtotal':True}, dataset=ds)

    build.save_excel('test.xlsx')
    wb = openpyxl.load_workbook('test.xlsx')
    os.remove('test.xlsx')

def test_add_table_bug(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_spss('tests/fixtures/Example Data (A).sav')

    build = tally.Build(name='client A', default_dataset=ds, table_of_contents=True)
    #build.logo = 'drive/MyDrive/projects/GWI_Tally/specs/gwi-logo.png'

    build.options.set_weight('weight_a')
    build.options.set_ci(['c%', 'counts'])
    build.options.set_stats(['mean', 'stddev'])
    build.options.set_hide_gridlines(2)

    sheet = build.add_sheet(banner=['gender'])

    sheet.options.set_base_position('outside')
    sheet.add_table(stub={'x' : "q1", 'base':'both'})
    sheet.add_table(stub={'x':'q2b', 'xtotal':True})

    sheet2 = build.add_sheet(banner=['gender'])

    sheet2.add_table(stub={'x' : "q1", 'ci':['c%','counts'], 'base':'both'})
    sheet2.add_table(stub={'x':'q2b', 'ci':['c%', 'counts'], 'xtotal':True, 'stats':['mean', 'stddev']}, dataset=ds)


    build.save_excel('test_simple_table_bug.xlsx')
    wb = openpyxl.load_workbook('test_simple_table_bug.xlsx')
    os.remove('test_simple_table_bug.xlsx')

def test_annotations(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')
    build = tally.Build(name='Client A', subtitle="Datasmoothie", default_dataset=ds)

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet.options.set_sig_test_levels(0.05)
    sheet.add_table(stub={'x':'q1', 'ci':['c%'], 'xtotal':True, 'stats':['mean']}, dataset=ds)

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet.options.set_annotations(True)
    sheet.options.set_sig_test_levels(0.05)
    sheet.add_table(stub={'x':'q1', 'ci':['c%'], 'xtotal':True, 'stats':['stddev']}, dataset=ds)

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet.add_table(stub={'x':'q1', 'ci':['c%'], 'xtotal':True, 'stats':['stddev']}, dataset=ds)

    build.save_excel('test_annotations.xlsx')

    wb = openpyxl.load_workbook('test_annotations.xlsx')
    os.remove('test_annotations.xlsx')


def test_default_options(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_spss('tests/fixtures/Example Data (A).sav')

    build = tally.Build(name='Client A', subtitle="Datasmoothie", default_dataset=ds)

    all_variables = ds.variables()
    banner_vars = ['gender', 'locality']
    stub_vars = [i for i in all_variables['single'] if i not in banner_vars]
    stub_vars = stub_vars + list(all_variables['delimited set'])

    sheet = build.add_sheet(banner=['gender', 'locality'])

    sheet.add_table(stub={'x':stub_vars[0]})

    build.sheets[0].add_table(stub={'x':'ethnicity'})

def test_table_without_formatting(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='Client A', subtitle="Datasmoothie", default_dataset=ds)

    sheet = build.add_sheet(banner=['gender', 'locality'])

    sheet.add_table(stub={'x' : 'q14r01c01'}) 
    sheet.add_table(stub={'x' : 'q14r02c01', 'stats':['mean']})
    sheet.add_table(stub={'x' : 'q14r03c01'})


    build.save_excel('test_table_without_formatting.xlsx')
    wb = openpyxl.load_workbook('test_table_without_formatting.xlsx')
    os.remove('test_table_without_formatting.xlsx')

def test_table_with_header_formatting(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='Client A', subtitle="Datasmoothie", default_dataset=ds)
    build.options.set_top_offset_after_header(1)

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet.options.set_base_position('outside')
    sheet.options.set_format('base', {'bold':True, 'border': 1, 'border_color':'000000'})
    sheet.options.set_base_labels("All respondents")

    sheet.add_table(stub={'x' : 'q14r01c01'}) 
    sheet.add_table(stub={'x' : 'q14r02c01', 'stats':['mean']})
    sheet.add_table(stub={'x' : 'q14r03c01'})


    build.save_excel('test_table_with_header_formatting.xlsx')
    os.remove('test_table_with_header_formatting.xlsx')


def test_table_with_totals(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='Client A', subtitle="Datasmoothie", default_dataset=ds)

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet.options.set_stub({'xtotal': True})
    sheet.options.set_column_format_for_type('base', 1, {"bold":True})
    sheet.options.set_column_format_for_type('percentage', 1, {"bold":True})

    sheet.add_table(stub={'x' : 'q14r01c01'}) 
    sheet.add_table(stub={'x' : 'q14r02c01', 'stats':['mean']})
    sheet.add_table(stub={'x' : 'q14r03c01'})

    build.save_excel('test_table_with_totals.xlsx')
    os.remove('test_table_with_totals.xlsx')


def test_add_simple_table(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='Client A', subtitle="Datasmoothie", default_dataset=ds)
    build.add_logo('tests/fixtures/datasmoothie-logo.png')

    build.options.set_sig_test_levels(0.05)
    build.options.set_weight('weight_b')
    sheet = build.add_sheet(banner=['gender > locality', 'locality > q2b'])

    sheet.options.set_banner_border(True)
    sheet.options.set_show_bases('both')

    sheet.options.set_show_table_base_column(True)
    sheet.options.set_column_format_for_type('base', 1, {"bold":True})
    sheet.options.set_column_format_for_type('percentage', 1, {"bold":True})

    sheet.options.set_base_position('outside')
    sheet.options.set_filter({'gender':[1], 'locality':[1,2]})

    sheet.add_table(stub={'x' : 'q14r01c01'}) 
    sheet.add_table(stub={'x' : 'q14r02c01', 'stats':['mean']})
    sheet.add_table(stub={'x' : 'q14r03c01'},
                          options={'row_format':{'rows':[4,5], 'format':{'bg_color':'F15A30'}}}
                   )

    build.save_excel('test_simple_table.xlsx')
    wb = openpyxl.load_workbook('test_simple_table.xlsx')
    os.remove('test_simple_table.xlsx')

def test_build_after_sheet_options(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')
    build = tally.Build(name='client A', default_dataset=ds)

    build.add_logo('tests/fixtures/datasmoothie-logo.png')
    build.options.freeze_panes(9,1)
    build.options.set_base_position('outside')
    build.options.set_answer_format('base', {"font_color":"F15A30", "bold":True,'text_wrap': True})

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet2 = build.add_sheet(banner=['gender', 'ethnicity'])
    sheet2.options.set_format('stats', {"font_color":"98B4DF"})
    sheet2.options.set_stats(stats=['mean', 'stddev'])

    sheet.add_table(stub={"x":'q1'})
    sheet2.add_table(stub={"x":'q2b'})

    build.options.set_format('base', {"bold":True})
    build.options.set_question_format('percentage', {"bold":True})

    build.save_excel('test_build_before_sheet.xlsx')
    wb = openpyxl.load_workbook('test_build_before_sheet.xlsx')
    os.remove('test_build_before_sheet.xlsx')

def test_global_options(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')
    build = tally.Build(name='client A', default_dataset=ds)

    build.add_logo('tests/fixtures/datasmoothie-logo.png')
    build.set_index_option('link_color', '57215B')
    build.font_size = '12'
    build.font_name = 'Arial'

    build.options.set_base_position('outside')
    build.options.set_format('base', {'bold':True, 'border': 1, 'border_color':'efefef'})
    build.options.set_weight('weight_a')

    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet2 = build.add_sheet(banner=['gender', 'locality'])

    [i.options.set_show_bases('both') for i in build.sheets]

    sheet.add_table(stub={'x' : 'q14r01c01', 'stats':["mean", "stddev"]}) 
    sheet.add_table(stub={'x' : 'q14r01c02', 'stats':["mean", "stddev"]}) 

    sheet2.add_table(stub={'x' : 'q14r02c01', 'stats':["mean", "stddev"]}) 
    sheet2.add_table(stub={'x' : 'q14r02c02', 'stats':["mean", "stddev"]}) 

    build.save_excel('test_global_options.xlsx')
    wb = openpyxl.load_workbook('test_global_options.xlsx')
    os.remove('test_global_options.xlsx')



def test_add_many_sheets(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='client A', default_dataset=ds)
    build.add_logo('tests/fixtures/datasmoothie-logo.png')

    questions = list(range(1,3))
    stores = list(range(1,4))

    for question in questions:
        sheet = build.add_sheet(banner=['gender', 'locality'])
        sheet.options.set_sig_test_levels([0.05])
        for store in stores:
            value = f"q14r0{question}c0{store}"
            sheet.add_table(stub={'x' : value} ,
                            options={'row_format':{'rows':[4,5], 
                                     'format':{'bg_color':'F15A30'}
                            }}
                            )
    build.save_excel('test_many_sheets.xlsx')
    wb = openpyxl.load_workbook('test_many_sheets.xlsx')
    os.remove('test_many_sheets.xlsx')

def test_build_options(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_quantipy('tests/fixtures/Example Data (A).json', 'tests/fixtures/Example Data (A).csv')

    build = tally.Build(name='client A', default_dataset=ds, table_of_contents=True)
    build.add_logo('tests/fixtures/datasmoothie-logo.png')

    build.options.freeze_panes(9,1)
    build.options.set_base_position('outside')
    build.options.set_answer_format('base', {"font_color":"F15A30", "bold":True,'text_wrap': True})
    build.options.set_format('base', {"bold":True})
    build.options.set_font("Comic Sans", 10)
    build.options.set_top_offset(3)
    build.options.freeze_panes(9,1)


    sheet = build.add_sheet(banner=['gender', 'locality'])
    sheet2 = build.add_sheet(banner=['@', 'gender', 'ethnicity'])
    sheet2.options.set_format('stats', {"font_color":"98B4DF"})
    sheet2.options.set_stats(stats=['mean', 'stddev'])

    sheet.add_table(stub={'x':'q1'})
    sheet.add_table(stub={'x':'q2b'})

    sheet2.add_table(stub={'x':'q1' })
    sheet2.add_table(stub={'x':'q2b'})

    build.save_excel('test_build_options.xlsx')
    wb = openpyxl.load_workbook('test_build_options.xlsx')
    os.remove('test_build_options.xlsx')

def test_add_change_base_label(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_spss('tests/fixtures/Example Data (A).sav')

    build = tally.Build(name='client A', default_dataset=ds, table_of_contents=True)

    build.options.set_hide_gridlines(2)
    build.options.set_base_position('outside')
    sheet = build.add_sheet(banner=['gender', 'locality'])
    build.options.set_weight('weight_a')

    sheet.options.set_base_labels('All GB', 'Unweighted base (All GB)')

    sheet.add_table(stub={'x':'q2b', 'xtotal':True, 'base':'both'})
    sheet.add_table(stub={'x':'q1', 'xtotal':True}, options={'base':'hide'})

    build.save_excel('test_add_change_base_label.xlsx')
    wb = openpyxl.load_workbook('test_add_change_base_label.xlsx')
    os.remove('test_add_change_base_label.xlsx')

def test_add_table_title(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_spss('tests/fixtures/Example Data (A).sav')

    build = tally.Build(name='client A', default_dataset=ds, table_of_contents=True)

    build.options.set_hide_gridlines(2)
    build.options.set_base_position('outside')
    sheet = build.add_sheet(banner=['gender', 'locality'])
    build.options.set_weight('weight_a')

    sheet.options.set_base_labels('All GB', 'Unweighted base (All GB)')

    sheet.add_table(stub={'x':'q2b', 'xtotal':True, 'base':'both'}, options={'title':{'text':'This is my title', 'format':{'bold':True, 'font_color':'#ebebeb'}}})
    sheet.add_table(stub={'x':'q1', 'xtotal':True}, options={'base':'hide'})

    build.save_excel('add_title.xlsx')
    wb = openpyxl.load_workbook('add_title.xlsx')
    os.remove('add_title.xlsx')


def test_add_slide(token, api_url, use_ssl):
    ds = tally.DataSet(api_key=token, host=api_url, ssl=use_ssl)
    ds.use_spss('tests/fixtures/Example Data (A).sav')

    build = tally.Build(name='client A', default_dataset=ds, table_of_contents=True)

    presentation = build.add_presentation('test', powerpoint_template='tests/fixtures/Datasmoothie_Template.pptx')

    presentation.add_slide(
        stub="q1", 
        banner="@",
        show='c%',
        options={
            'chart_type':'column_clustered',
            'title':'What is your main sporting activity?',
            'data_labels': True,
            'data_labels_position':'outside_end'
        }
    )
    presentation.add_slide(
        stub="q1", 
        banner="q2b",
        show='c%',
        options={
            'chart_type':'column_clustered',
            'title':'What is your main activity and how frequently do you exercise?'
        }
    )
    presentation.add_slide(
        stub="q14r01c01", 
        banner="Wave",
        show='mean',
        options={
            'chart_type':'line',
            'template':1,
            'title':'I hade a good experience in the store',
            'value_axis_minimum_scale':0,
            'value_axis_maximum_scale':5           
        }
    )

    presentation.add_slide(
        stub="q1", 
        banner="@",
        show='c%',
        table='counts',
        options={
            'chart_type':'pie',
            'template':2,
            'title':'What sports do you do?'
        }
    )

    presentation.add_slide(
        stub="q1", 
        banner="gender",
        show='r%',
        options={
            'chart_type':'column_clustered',
            'template':1,
            'title':'What is your main sporting activity?'
        }
    )

    presentation.save_powerpoint('test.pptx')