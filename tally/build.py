import pandas as pd
import json
import copy
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .build_defaults import build_default_formats
from .options import Options
from .sheet import Sheet
from .presentation import Presentation

class Build:
    """ Represents crosstabs and settings to build a deliverable, such as Excel tables.



    Parameters
    ----------
      name: string
        Name for the dataset
    """

    def __init__(self, name=None, subtitle=None, default_dataset=None, table_of_contents=False, logo=None):
        self.name = name
        self.subtitle = subtitle
        self.default_dataset = default_dataset
        self.table_of_contents = table_of_contents
        self.sheets = []
        self.logo = logo
        self.font_name = None
        self.font_size = None
        self.index_options = {
            'header_color': 'ffffff',
            'body_color': 'ffffff',
            'link_color': '2A64C5'
        }
        self.options = Options(
            formats=copy.deepcopy(build_default_formats),
            parent=self
        )

    def add_presentation(self, name=None, powerpoint_template=None):
        if name is None:
            name = "Presentation {}" #.format(str(len(self.sheets)+1))
        presentation = Presentation(
            default_dataset=self.default_dataset, 
            name=name,
            powerpoint_template=powerpoint_template,
            parent=self
            )
        return presentation

    def show(self, sheet=None, table=None) -> pd.DataFrame:
        """Return the dataframe or dataframes specified by the params.
        Keyword arguments:
        sheet -- the sheet number (default None)
        table -- the table number (default None)
        """
        if sheet is None:
            raise TypeError("sheet argument missing")
        if not isinstance(sheet, int):
            raise ValueError(f"The parameter sheet={sheet} must be an integer")
        if table is not None and not isinstance(table, int):
            raise ValueError(f"The parameter table={table} must be an integer")

        if table is None:
            df_list = []
            for t in self.sheets[sheet].tables:
                df_list.append(t['dataframe'])
            df = pd.concat(df_list)
        else:
            df = self.sheets[sheet].tables[table]['dataframe']
        if 'FORMAT' in df.columns:
            df = df.drop("FORMAT", axis=1, level=0)
        return df

    def add_sheet(self, name=None, banner='@'):
        """ Add sheet to a build

        The sheet can then be used both to change settings via its options and to store a list of tables.

        Parameters
        ----------

        name: string
            Name of the sheet. Appears in the table of contents.
        banner: list of strings
            List of variables to display across the top of the sheet.


        Example
        -------
            build.add_sheet(name='Q1 - gender, location', banner=['gender', 'location'])
        """
        if name is None:
            name = "Table {}".format(str(len(self.sheets)+1))
        sheet = Sheet(
            banner=banner, 
            default_dataset=self.default_dataset, 
            name=name,
            parent=self
            )
        self.sheets.append(sheet)
        return sheet

    def add_logo(self, path):
        self.logo = path

    def set_index_option(self, name, option):
        self.index_options[name] = option

    def save_excel(self, filename):
        dataframes_payload = []
        formats = []
        for sheet in self.sheets:
            df = sheet.build_dataframes(self.options)
            df = df.replace({'null':0})
            payload = json.loads(df.to_json(orient='split'))
            payload['index_names'] = df.index.names.copy()
            payload['column_names'] = df.columns.names.copy()
            dataframes_payload.append(payload)        
            formats.append(sheet.options.formats)

        self.default_dataset.build_excel_from_dataframes(filename=filename, dataframes=dataframes_payload, client_formats=formats)


        # wrapped in try because openpyxl is brittle
        try:
            font_name = self.options.global_options['font_name']
            font_size = self.options.global_options['font_size']

            heights = {
                'Question': 30,
                'Values': 70,
                'Test-IDs': 30
            }

            wb = load_workbook(filename)

            
            # add table of contents
            if len(self.sheets)>1:
                offset_y = 12

                wb_sheet = wb.create_sheet('Contents', 0)
                wb_sheet.cell(row=offset_y, column=2, value='Contents')
                wb_sheet.cell(row=offset_y, column=2).font = Font(name=font_name, size=font_size)
                wb_sheet.column_dimensions['B'].width = 20
                wb_sheet.column_dimensions['C'].width = 70

                if self.name is not None:
                    title = wb_sheet.cell(7, 2, value=self.name)
                    title.font = Font(size="20", name=font_name)
                if self.subtitle is not None:
                    subtitle = wb_sheet.cell(8, 2, value=self.subtitle)
                    subtitle.font = Font(size="14", name=font_name)

                for index, sheet in enumerate(self.sheets):
                    wb_sheet.cell(row=index+1+offset_y, column=3, value=sheet.get_name())
                    wb_sheet.cell(row=index+1+offset_y, column=3).font = Font(name=font_name, size=font_size)
                    link_value = '=HYPERLINK("#{}!A1", "Table {}")'.format(wb.sheetnames[index+1], index+1)
                    cell = wb_sheet.cell(row=index+1+offset_y, column=2, value=link_value)
                    cell.font = Font(color="2A64C5", underline="single", name=font_name, size=font_size)

                if self.logo is not None:
                    img = Image(self.logo)
                    wb_sheet.add_image(img, 'B2')

                for row_range in range(1, 10):
                    for col_range in range(1, 30):
                        color_cell = wb_sheet.cell(row_range, col_range)
                        color_cell.fill = PatternFill(start_color=self.index_options['header_color'], end_color=self.index_options['header_color'], fill_type="solid")

                # self.sheets tells us how many entries are in the index
                for row_range in range(10, len(self.sheets)+100):
                    for col_range in range(1, 30):
                        color_cell = wb_sheet.cell(row_range, col_range)
                        color_cell.fill = PatternFill(start_color=self.index_options['body_color'], end_color=self.index_options['body_color'], fill_type="solid")
                        #color_cell.font = default_font

                for index, sh in enumerate(self.sheets):
                    top_offset = sh.options.formats['offsets']['top']
                    wb_sheet = wb.worksheets[index+1]
                    wb_sheet['A3'].font = Font(bold=True, size=16, name=font_name)
                    if wb_sheet.cell(top_offset+1, 2).value == 'Total':
                        wb_sheet.merge_cells(
                            start_row=top_offset+1, 
                            end_row=top_offset+len(self.sheets[0].tables[0]['dataframe'].columns.names), 
                            start_column=2, 
                            end_column=2
                        )

                    for index, header in enumerate(sh.tables[0]['dataframe'].columns.names):
                        wb_sheet.row_dimensions[top_offset+index+1].height=heights[header]
                        for col_range in range(1,100):
                            wb_sheet.cell(top_offset+index+1,col_range).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                            wb_sheet.cell(top_offset+index+1,col_range).font = Font(
                                name=font_name,
                                size=font_size,
                                bold=True
                            )




            else:
                wb_sheet = wb.worksheets[0]
                if self.logo is not None:                
                    img = Image(self.logo)
                    wb_sheet.add_image(img, 'B2')

                for row_range in range(1, self.sheets[0].options.formats['offsets']['top']+1):
                    for col_range in range(1, 30):
                        color_cell = wb_sheet.cell(row_range, col_range)
                        color_cell.fill = PatternFill(start_color=self.index_options['header_color'], end_color=self.index_options['body_color'], fill_type="solid")

                top_offset = self.sheets[0].options.formats['offsets']['top']
                if wb_sheet.cell(top_offset+1, 2).value == 'Total':
                    wb_sheet.merge_cells(
                        start_row=top_offset+1, 
                        end_row=top_offset+len(self.sheets[0].tables[0]['dataframe'].columns.names), 
                        start_column=2, 
                        end_column=2
                    )

                # if we have the base outside, remove the first base line from the top
                if 'base' in self.options.table_options:
                    base_option = self.options.table_options['base']
                elif 'base' in sheet.options.table_options:
                    base_option = sheet.options.table_options['base']
                else:
                    base_option = None
                if base_option == 'outside':
                    row_for_removal = top_offset + len(self.sheets[0].tables[0]['dataframe'].columns.names) + 1
                    wb_sheet.delete_rows(row_for_removal)
                top_offset + len(self.sheets[0].tables[0]['dataframe'].columns.names) + 1
                for index, header in enumerate(self.sheets[0].tables[0]['dataframe'].columns.names):
                    wb_sheet.row_dimensions[top_offset+index+1].height=heights[header]
                    for col_range in range(1,100):
                        wb_sheet.cell(top_offset+index+1,col_range).alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                        wb_sheet.cell(top_offset+index+1,col_range).font = Font(
                            name=font_name,
                            size=font_size,
                            bold=True
                        )

            wb.save(filename)
        except Exception as e:
            raise Exception('Error: Something went wrong in prettifying the Excel. Make sure all colors are defined as ffffff rather than #ffffff.')

    def table_count(self):
        table_counts = [i.table_count() for i in self.sheets]
        return sum(table_counts)
